
import numpy

import simtk.unit as units

import openpyxl # Excel spreadsheet I/O (for Auto iTC-200)

class ITCProtocol(object):
    def __init__(self, name, sample_prep_method, itc_method, analysis_method):
        """
        Parameters
        ----------
        name : str
           The name of the protocol.
        sample_prep_method : str
           The name of the 'SamplePrepMethod' to be written to the Excel file for the Auto iTC-200.
        itc_method : str
           The name of the 'ItcMethod' to be written to the Excel file for the Auto iTC-200.
        analysis_method : str
           The name of the 'AnalysisMethod' to be written to the Excel file for the Auto iTC-200.

        """
        self.name = name
        self.sample_prep_method = sample_prep_method
        self.itc_method = itc_method
        self.analysis_method = analysis_method

class ITCExperiment(object):
    def __init__(self, name, syringe_source, cell_source, protocol, buffer_source=None, syringe_concentration=None, cell_concentration=None):
        """
        Parameters
        ----------
        name : str
           Name of the ITC experiment.
        syringe_source : Solution
           Source for syringe solution.
        cell_source : Solution
           Source for cell solution.
        protocol : ITCProtocol
           Protocol to be used for ITC experiment and analysis.
        buffer_source : Labware
           Source for buffer.
        syringe_concentration : simtk.unit.Quantity with units compatible with moles/liter, optional, default=None
           If specified, syringe source will be diluted to specified concentration.
           Buffer source must be specified.
        cell_concentration : simtk.unit.Quantity with units compatible with moles/liter, optional, default=None
           If specified, cell source will be diluted to specified concentration.
           Buffer source must be specified.

        WARNING
        -------
        Do not change class member fields after initialization.  Dilution factors are not recomputed.

        """
        self.name = name
        self.syringe_source = syringe_source
        self.cell_source = cell_source
        self.protocol = protocol

        # Store data.
        self.buffer_source = buffer_source
        self.syringe_concentration = syringe_concentration
        self.cell_concentration = cell_concentration

        # Compute dilution factors.
        self.syringe_dilution_factor = None
        self.cell_dilution_factor = None

        if syringe_concentration is not None:
            self.syringe_dilution_factor = syringe_concentration / syringe_source.concentration
            self.syringe_concentration= syringe_concentration

        if cell_concentration is not None:
            self.cell_dilution_factor = cell_concentration / cell_source.concentration
            self.cell_concentration = cell_concentration

        # If dilution is required, make sure buffer source is specified.
        if (self.syringe_dilution_factor is not None):
            if (buffer_source is None):
                raise Exception("buffer must be specified if either syringe or cell concentrations are specified")
            if (self.syringe_dilution_factor > 1.0):
                raise Exception("Requested syringe concentration (%s) is greater than syringe source concentration (%s)." % (str(syringe_concentration), str(syringe_source.concentration)))

        if (self.cell_dilution_factor is not None):
            if (buffer_source is None):
                raise Exception("buffer must be specified if either syringe or cell concentrations are specified")
            if (self.cell_dilution_factor > 1.0):
                raise Exception("Requested cell concentration (%s) is greater than cell source concentration (%s)." % (str(cell_concentration), str(cell_source.concentration)))

class ITCExperimentSet(object):
    def __init__(self, name):
        """
        Parameters
        ----------
        name : str
           Name of the experiment set.

        """
        self.name = name # name of the experiment
        self.experiments = list() # list of experiments to set up
        self.destination_plates = list() # ITC plates available for use in experiment

        self._validated = False

    def addDestinationPlate(self, plate):
        """
        Add the specified destination plate to the plate set usable for setting up ITC experiments.

        Parameters
        ----------
        plate : Labware
           The empty ITC destination plate to add to the experiment set.

        """

        # TODO: Check if specified plate is allowed type of labware for use in Auto iTC-200.
        self.destination_plates.append(plate)

    def addExperiment(self, experiment):
        """
        Add the specified ITC experiment to the experiment set.

        Parameters
        ----------
        experiment : ITCExperiment
           The ITC experiment to add.

        """
        self.experiments.append(experiment)

    def _wellIndexToName(self, index):
        """
        Return the 96-well name (e.g. 'A6', 'B7') corresponding to Tecan well index.

        Parameters
        ----------
        index : int
           Tecan well index (back to front, left to right), numbered from 1..96

        Returns
        -------
        well_name : str
           Well name for ITC plate (e.g. 'A6'), numbered from A1 to H12

        """
        row = int((index-1) % 8)
        column = int((index-1) / 8)
        rownames = 'ABCDEFGH'
        well_name = rownames[row] + '%d'%(column+1)
        return well_name

    class ITCData(object):
        def __init__(self):
            fieldnames = ['DataFile', 'SampleName', 'SamplePrepMethod', 'ItcMethod', 'AnalysisMethod', 'CellConcentration', 'PipetteConcentration', 'CellSource', 'PipetteSource', 'PreRinseSource', 'SaveSampleDestination']
            for fieldname in fieldnames:
                setattr(self, fieldname, None)

    class TecanData(object):
        def __init__(self):
            fieldnames = ['cell_destination', 'cell_platename', 'cell_wellindex', 'syringe_plateindex', 'syringe_platename', 'syringe_wellindex']
            for fieldname in fieldnames:
                setattr(self, fieldname, None)

    def _resetTrackedQuantities(self):
        self._tracked_quantities = dict()

    def _trackQuantities(self, thing, volume):
        try:
            name = thing.name
        except:
            name = thing.RackLabel

        #print name, volume

        if name in self._tracked_quantities:
            self._tracked_quantities[name] += volume
        else:
            self._tracked_quantities[name] = volume

    def validate(self):
        """
        Validate that the specified set of ITC experiments can actually be set up, raising an exception if not.

        Additional experiment data fields (tecandata, itcdata)

        """

        # TODO: Try to set up experiment, throwing exception upon failure.

        # Make a list of all the possible destination pipetting locations.
        from automation import PipettingLocation
        # TODO: Change this to go left-to-right in ITC plates?
        destination_locations = list()
        for (plate_index, plate) in enumerate(self.destination_plates):
            PlateNumber = plate_index + 1
            for index in range(96):
                Position = index + 1
                WellName = self._wellIndexToName(Position)
                location = PipettingLocation(plate.RackLabel, plate.RackType, Position)
                # Add plate number and well name for Auto iTC-200.
                location.PlateNumber = PlateNumber
                location.WellName = WellName
                destination_locations.append(location)

        # Build worklist script.
        worklist_script = ""

        # Reset tracked quantities.
        self._resetTrackedQuantities()

        for (experiment_number, experiment) in enumerate(self.experiments):
            experiment.experiment_number = experiment_number

            itcdata = ITCExperimentSet.ITCData()
            tecandata = ITCExperimentSet.TecanData()

            # Find a place to put cell contents.
            if len(destination_locations) == 0:
                raise Exception("Ran out of destination plates for experiment %d / %d" % (experiment_number, len(self.experiments)))
            tecandata.cell_destination = destination_locations.pop(0)

            cell_volume = 400.0 # microliters
            transfer_volume = cell_volume

            if (experiment.cell_dilution_factor is not None):
                # Compute buffer volume needed.
                buffer_volume = cell_volume * (1.0 - experiment.cell_dilution_factor)
                transfer_volume = cell_volume - buffer_volume 

                # Schedule buffer transfer.
                tipmask = 1
                worklist_script += 'A;%s;;%s;%d;;%f;;;%d\r\n' % (experiment.buffer_source.RackLabel, experiment.buffer_source.RackType, 1, buffer_volume, tipmask)
                worklist_script += 'D;%s;;%s;%d;;%f;;;%d\r\n' % (tecandata.cell_destination.RackLabel, tecandata.cell_destination.RackType, tecandata.cell_destination.Position, buffer_volume, tipmask)
                worklist_script += 'W;\r\n' # queue wash tips
                self._trackQuantities(experiment.buffer_source, buffer_volume * units.microliters)

            # Schedule cell solution transfer.
            tipmask = 2
            try:
                # Assume source is Solution.
                worklist_script += 'A;%s;;%s;%d;;%f;;;%d\r\n' % (experiment.cell_source.location.RackLabel, experiment.cell_source.location.RackType, experiment.cell_source.location.Position, transfer_volume, tipmask)
            except:
                # Assume source is Labware.
                worklist_script += 'A;%s;;%s;%d;;%f;;;%d\r\n' % (experiment.cell_source.RackLabel, experiment.cell_source.RackType, 2, transfer_volume, tipmask)
            worklist_script += 'D;%s;;%s;%d;;%f;;;%d\r\n' % (tecandata.cell_destination.RackLabel, tecandata.cell_destination.RackType, tecandata.cell_destination.Position, transfer_volume, tipmask)
            worklist_script += 'W;\r\n' # queue wash tips
            self._trackQuantities(experiment.cell_source, transfer_volume * units.microliters)

            # Find a place to put syringe contents.
            if len(destination_locations) == 0:
                raise Exception("Ran out of destination plates for experiment %d / %d" % (experiment_number, len(self.experiments)))
            tecandata.syringe_destination = destination_locations.pop(0)

            syringe_volume = 120.0 # microliters
            transfer_volume = cell_volume

            if (experiment.syringe_dilution_factor is not None):
                # Compute buffer volume needed.
                buffer_volume = syringe_volume * (1.0 - experiment.syringe_dilution_factor)
                transfer_volume = syringe_volume - buffer_volume 

                # Schedule buffer transfer.
                tipmask = 4
                worklist_script += 'A;%s;;%s;%d;;%f;;;%d\r\n' % (experiment.buffer_source.RackLabel, experiment.buffer_source.RackType, 3, buffer_volume, tipmask)
                worklist_script += 'D;%s;;%s;%d;;%f;;;%d\r\n' % (tecandata.syringe_destination.RackLabel, tecandata.syringe_destination.RackType, tecandata.syringe_destination.Position, buffer_volume, tipmask)
                worklist_script += 'W;\r\n' # queue wash tips
                self._trackQuantities(experiment.buffer_source, buffer_volume * units.microliters)

            # Schedule syringe solution transfer.
            tipmask = 8
            try:
                # Assume source is Solution.
                worklist_script += 'A;%s;;%s;%d;;%f;;;%d\r\n' % (experiment.syringe_source.location.RackLabel, experiment.syringe_source.location.RackType, experiment.syringe_source.location.Position, transfer_volume, tipmask)
            except:
                # Assume source is Labware.
                worklist_script += 'A;%s;;%s;%d;;%f;;;%d\r\n' % (experiment.syringe_source.RackLabel, experiment.syringe_source.RackType, 4, transfer_volume, tipmask)
            worklist_script += 'D;%s;;%s;%d;;%f;;;%d\r\n' % (tecandata.syringe_destination.RackLabel, tecandata.syringe_destination.RackType, tecandata.syringe_destination.Position, transfer_volume, tipmask)
            worklist_script += 'W;\r\n' # queue wash tips
            self._trackQuantities(experiment.syringe_source, transfer_volume * units.microliters)

            # Finish worklist section.
            worklist_script += 'B;\r\n' # execute queued batch of commands

            # Create datafile name.
            datecode = '20140530' # TODO: Build with date
            seriescode = 'a' # TODO: Use intelligent coding?
            indexcode = '%d' % (experiment_number + 1)
            itcdata.DataFile = datecode + seriescode + indexcode

            itcdata.SampleName = experiment.name
            itcdata.SamplePrepMethod = experiment.protocol.sample_prep_method
            itcdata.ItcMethod = experiment.protocol.itc_method
            itcdata.AnalysisMethod = experiment.protocol.analysis_method

            millimolar = 0.001 * units.moles / units.liter
            try:
                itcdata.CellConcentration = experiment.cell_source.concentration / millimolar
            except:
                itcdata.CellConcentration = 0

            try:
                itcdata.PipetteConcentration = experiment.pipette.concentration / millimolar
            except:
                itcdata.PipetteConcentration = 0

            itcdata.CellSource = 'Plate%d, %s' % (tecandata.cell_destination.PlateNumber, tecandata.cell_destination.WellName)
            itcdata.PipetteSource = 'Plate%d, %s' % (tecandata.syringe_destination.PlateNumber, tecandata.syringe_destination.WellName)

            # TODO: Autodetect if prerinse is used.
            itcdata.PreRinseSource = ''

            # TODO: Autodetect if sample destination is used.
            itcdata.SaveSampleDestination = itcdata.CellSource

            # Store Tecan and Excel data for this experiment.
            experiment.tecandata = tecandata
            experiment.itcdata = itcdata

        # Save Tecan worklist.
        self.worklist = worklist_script

        # Report tracked quantities.
        print "Necessary volumes:"
        keys = self._tracked_quantities.keys()
        keys.sort()
        for key in keys:
            print "%32s %12.3f mL" % (key, self._tracked_quantities[key] / units.milliliters)

        # Set validated flag.
        self._validated = True

    def writeTecanWorklist(self, filename):
        """
        Write the Tecan worklist for the specified experiment set.

        Parameters
        ----------
        filename : str
           The name of the Tecan worklist file to write.

        """
        if not self._validated:
            self.validate()

        outfile = open(filename, 'w')
        outfile.write(self.worklist)
        outfile.close()

    def writeAutoITCExcel(self, filename):
        """
        Write the Excel file for the specified experiment set to be loaded into the Auto iTC-200.

        Parameters
        ----------
        filename : str
           The name of the Excel file to write.

        """

        if not self._validated:
            self.validate()

        # Create new Excel spreadsheet.
        from openpyxl import Workbook
        wb = Workbook()

        # Create plate sheet.
        ws = wb.get_active_sheet()
        ws.title = 'plate'

        # Create header.
        row = 0
        fieldnames = ['DataFile', 'SampleName', 'SamplePrepMethod', 'ItcMethod', 'AnalysisMethod', 'CellConcentration', 'PipetteConcentration', 'CellSource', 'PipetteSource', 'PreRinseSource', 'SaveSampleDestination']
        for (column, fieldname) in enumerate(fieldnames):
            ws.cell(row=row, column=column).value = fieldname

        # Create experiments.
        for experiment in self.experiments:
            row += 1
            for (column, fieldname) in enumerate(fieldnames):
                ws.cell(row=row, column=column).value = getattr(experiment.itcdata, fieldname)

        # Write workbook.
        wb.save(filename)


